from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, Protocol

from openpyxl import load_workbook

from market_info.ai.extractor import ProjectExtractor
from market_info.ai.schemas import ExtractedProject
from market_info.config import Settings
from market_info.dedupe.normalizers import normalize_project_name


ARTICLE_HEADERS = [
    "article_id",
    "account_name",
    "title",
    "published_at",
    "url",
    "body_path",
    "is_project_article",
    "expected_project_count",
    "notes",
]

EXPECTED_PROJECT_HEADERS = [
    "article_id",
    "project_group_id",
    "project_name",
    "project_info",
    "province",
    "city",
    "detailed_address",
    "company_name",
    "investment_yi",
    "industry",
    "field",
    "market",
    "status",
    "evidence",
    "notes",
]

EXPECTED_DEDUPE_HEADERS = [
    "project_group_id",
    "article_id",
    "project_name",
    "expected_decision",
    "expected_status_change",
    "status_order",
    "notes",
]


class ExtractorProtocol(Protocol):
    def extract(self, title: str, text: str) -> list[ExtractedProject]:
        ...


@dataclass(frozen=True)
class GoldenArticle:
    article_id: str
    account_name: str
    title: str
    published_at: str
    url: str
    body_path: str
    body_text: str
    is_project_article: bool
    expected_project_count: int
    notes: str


@dataclass(frozen=True)
class ExpectedProject:
    article_id: str
    project_group_id: str
    project_name: str
    project_info: str
    province: str
    city: str
    detailed_address: str
    company_name: str
    investment_yi: float | None
    industry: str
    field: str
    market: str
    status: str
    evidence: str
    notes: str


@dataclass(frozen=True)
class ExpectedDedupe:
    project_group_id: str
    article_id: str
    project_name: str
    expected_decision: str
    expected_status_change: bool
    status_order: int | None
    notes: str


@dataclass(frozen=True)
class GoldenLabels:
    labels_path: Path
    articles: list[GoldenArticle]
    expected_projects: list[ExpectedProject]
    expected_dedupe: list[ExpectedDedupe]


@dataclass(frozen=True)
class PredictedProject:
    article_id: str
    project: ExtractedProject


@dataclass(frozen=True)
class DedupePrediction:
    project_group_id: str | None
    article_id: str
    project_name: str
    decision: str
    status_change: bool


@dataclass(frozen=True)
class ExtractionMetrics:
    project_precision: float
    project_recall: float
    field_accuracy: float
    status_accuracy: float
    investment_accuracy: float
    hallucination_count: int
    missed_count: int


@dataclass(frozen=True)
class DedupeMetrics:
    dedupe_accuracy: float
    false_merge_count: int
    missed_merge_count: int
    status_change_accuracy: float


@dataclass(frozen=True)
class ErrorSamples:
    missed_articles: list[str]
    hallucinated_projects: list[dict[str, object]]
    field_errors: list[dict[str, object]]
    false_merges: list[dict[str, object]]
    missed_merges: list[dict[str, object]]


@dataclass(frozen=True)
class EvaluationReport:
    extraction: ExtractionMetrics
    dedupe: DedupeMetrics
    error_samples: ErrorSamples

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


DedupeDecider = Callable[[list[PredictedProject]], list[DedupePrediction]]


def load_golden_labels(labels_path: Path) -> GoldenLabels:
    workbook = load_workbook(labels_path, data_only=True)
    _validate_sheet(workbook, "articles", ARTICLE_HEADERS)
    _validate_sheet(workbook, "expected_projects", EXPECTED_PROJECT_HEADERS)
    _validate_sheet(workbook, "expected_dedupe", EXPECTED_DEDUPE_HEADERS)
    labels_dir = labels_path.parent
    articles = [
        _article_from_row(row, labels_dir)
        for row in _rows_as_dicts(workbook["articles"])
    ]
    expected_projects = [
        _expected_project_from_row(row)
        for row in _rows_as_dicts(workbook["expected_projects"])
    ]
    expected_dedupe = [
        _expected_dedupe_from_row(row)
        for row in _rows_as_dicts(workbook["expected_dedupe"])
    ]
    _validate_label_consistency(articles, expected_projects)
    return GoldenLabels(
        labels_path=labels_path,
        articles=articles,
        expected_projects=expected_projects,
        expected_dedupe=expected_dedupe,
    )


def evaluate_golden(
    labels_path: Path,
    extractor: ExtractorProtocol | None = None,
    dedupe_decider: DedupeDecider | None = None,
    report_path: Path | None = None,
) -> EvaluationReport:
    labels = load_golden_labels(labels_path)
    extractor = extractor or _default_extractor()

    predictions: list[PredictedProject] = []
    for article in labels.articles:
        for project in extractor.extract(article.title, article.body_text):
            predictions.append(PredictedProject(article.article_id, project))

    matched_pairs, missed_projects, hallucinated_projects = _match_projects(
        labels.expected_projects,
        predictions,
    )
    extraction_metrics, extraction_errors = _calculate_extraction_metrics(
        matched_pairs,
        missed_projects,
        hallucinated_projects,
        len(labels.expected_projects),
        len(predictions),
    )

    dedupe_predictions = (
        dedupe_decider(predictions)
        if dedupe_decider is not None
        else _default_dedupe_decider(predictions)
    )
    dedupe_metrics, dedupe_errors = _calculate_dedupe_metrics(
        labels.expected_dedupe,
        dedupe_predictions,
    )

    report = EvaluationReport(
        extraction=extraction_metrics,
        dedupe=dedupe_metrics,
        error_samples=ErrorSamples(
            missed_articles=sorted({project.article_id for project in missed_projects}),
            hallucinated_projects=extraction_errors["hallucinated_projects"],
            field_errors=extraction_errors["field_errors"],
            false_merges=dedupe_errors["false_merges"],
            missed_merges=dedupe_errors["missed_merges"],
        ),
    )
    output_path = report_path or labels_path.parent / "evaluation_report.json"
    output_path.write_text(
        json.dumps(report.to_dict(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return report


def _default_extractor() -> ProjectExtractor:
    settings = Settings()
    return ProjectExtractor(
        settings.ai_base_url,
        settings.ai_api_key,
        settings.ai_extraction_model,
        timeout=getattr(settings, "ai_extraction_timeout_seconds", 180),
    )


def _validate_sheet(workbook, sheet_name: str, expected_headers: list[str]) -> None:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Golden labels sheet '{sheet_name}' is missing; "
            f"expected headers: {expected_headers}"
        )

    actual_headers = [
        cell.value
        for cell in next(workbook[sheet_name].iter_rows(min_row=1, max_row=1))
    ]
    if actual_headers != expected_headers:
        raise ValueError(
            f"Golden labels sheet '{sheet_name}' has invalid headers; "
            f"expected headers: {expected_headers}; actual headers: {actual_headers}"
        )


def _validate_label_consistency(
    articles: list[GoldenArticle],
    expected_projects: list[ExpectedProject],
) -> None:
    project_count_by_article_id: dict[str, int] = {}
    for project in expected_projects:
        project_count_by_article_id[project.article_id] = (
            project_count_by_article_id.get(project.article_id, 0) + 1
        )

    article_ids = {article.article_id for article in articles}
    for project_article_id in project_count_by_article_id:
        if project_article_id not in article_ids:
            raise ValueError(
                f"article_id {project_article_id} appears in expected_projects "
                "but not in articles."
            )

    for article in articles:
        actual_count = project_count_by_article_id.get(article.article_id, 0)
        if not article.is_project_article and actual_count:
            raise ValueError(
                f"article_id {article.article_id} has is_project_article is False "
                "but expected_projects contains rows."
            )
        if article.is_project_article and article.expected_project_count <= 0:
            raise ValueError(
                f"article_id {article.article_id} has is_project_article is True "
                "but expected_project_count is not greater than 0."
            )
        if article.expected_project_count != actual_count:
            raise ValueError(
                f"article_id {article.article_id} expected_project_count "
                f"is {article.expected_project_count}, but expected_projects "
                f"contains {actual_count} rows."
            )


def _rows_as_dicts(sheet) -> list[dict[str, object]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    return [
        {
            header: value
            for header, value in zip(headers, row)
        }
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def _article_from_row(row: dict[str, object], labels_dir: Path) -> GoldenArticle:
    body_path = _text(row.get("body_path"))
    body_file = labels_dir / body_path
    return GoldenArticle(
        article_id=_text(row.get("article_id")),
        account_name=_text(row.get("account_name")),
        title=_text(row.get("title")),
        published_at=_text(row.get("published_at")),
        url=_text(row.get("url")),
        body_path=body_path,
        body_text=body_file.read_text(encoding="utf-8"),
        is_project_article=_bool(row.get("is_project_article")),
        expected_project_count=_int(row.get("expected_project_count")) or 0,
        notes=_text(row.get("notes")),
    )


def _expected_project_from_row(row: dict[str, object]) -> ExpectedProject:
    return ExpectedProject(
        article_id=_text(row.get("article_id")),
        project_group_id=_text(row.get("project_group_id")),
        project_name=_text(row.get("project_name")),
        project_info=_text(row.get("project_info")),
        province=_text(row.get("province")),
        city=_text(row.get("city")),
        detailed_address=_text(row.get("detailed_address")),
        company_name=_text(row.get("company_name")),
        investment_yi=_float(row.get("investment_yi")),
        industry=_text(row.get("industry")),
        field=_text(row.get("field")),
        market=_text(row.get("market")),
        status=_text(row.get("status")),
        evidence=_text(row.get("evidence")),
        notes=_text(row.get("notes")),
    )


def _expected_dedupe_from_row(row: dict[str, object]) -> ExpectedDedupe:
    return ExpectedDedupe(
        project_group_id=_text(row.get("project_group_id")),
        article_id=_text(row.get("article_id")),
        project_name=_text(row.get("project_name")),
        expected_decision=_text(row.get("expected_decision")),
        expected_status_change=_bool(row.get("expected_status_change")),
        status_order=_int(row.get("status_order")),
        notes=_text(row.get("notes")),
    )


def _match_projects(
    expected_projects: list[ExpectedProject],
    predictions: list[PredictedProject],
) -> tuple[
    list[tuple[ExpectedProject, PredictedProject]],
    list[ExpectedProject],
    list[PredictedProject],
]:
    unmatched_predictions = list(predictions)
    matched_pairs: list[tuple[ExpectedProject, PredictedProject]] = []
    missed_projects: list[ExpectedProject] = []

    for expected in expected_projects:
        match_index = next(
            (
                index
                for index, prediction in enumerate(unmatched_predictions)
                if _projects_match(expected, prediction)
            ),
            None,
        )
        if match_index is None:
            missed_projects.append(expected)
            continue
        matched_pairs.append((expected, unmatched_predictions.pop(match_index)))

    return matched_pairs, missed_projects, unmatched_predictions


def _projects_match(expected: ExpectedProject, prediction: PredictedProject) -> bool:
    return (
        expected.article_id == prediction.article_id
        and normalize_project_name(expected.project_name)
        == normalize_project_name(prediction.project.project_name)
    )


def _calculate_extraction_metrics(
    matched_pairs: list[tuple[ExpectedProject, PredictedProject]],
    missed_projects: list[ExpectedProject],
    hallucinated_projects: list[PredictedProject],
    expected_count: int,
    prediction_count: int,
) -> tuple[ExtractionMetrics, dict[str, list[dict[str, object]]]]:
    field_correct = 0
    field_total = 0
    status_correct = 0
    status_total = 0
    investment_correct = 0
    investment_total = 0
    field_errors: list[dict[str, object]] = []

    for expected, prediction in matched_pairs:
        for expected_name, predicted_name in _FIELD_PAIRS:
            expected_value = getattr(expected, expected_name)
            if expected_value in (None, ""):
                continue
            predicted_value = getattr(prediction.project, predicted_name)
            field_total += 1
            if _values_equal(expected_value, predicted_value):
                field_correct += 1
            else:
                field_errors.append(
                    {
                        "article_id": expected.article_id,
                        "project_name": expected.project_name,
                        "field": expected_name,
                        "expected": expected_value,
                        "actual": predicted_value,
                    }
                )

        if expected.status:
            status_total += 1
            status_correct += int(expected.status == prediction.project.status)

        if expected.investment_yi is not None:
            investment_total += 1
            investment_correct += int(
                _numbers_close(
                    expected.investment_yi,
                    prediction.project.investment_amount_yi,
                )
            )

    hallucinated_payload = [
        {
            "article_id": prediction.article_id,
            "project_name": prediction.project.project_name,
        }
        for prediction in hallucinated_projects
    ]
    matched_count = len(matched_pairs)
    return (
        ExtractionMetrics(
            project_precision=_ratio(matched_count, prediction_count),
            project_recall=_ratio(matched_count, expected_count),
            field_accuracy=_ratio(field_correct, field_total),
            status_accuracy=_ratio(status_correct, status_total),
            investment_accuracy=_ratio(investment_correct, investment_total),
            hallucination_count=len(hallucinated_projects),
            missed_count=len(missed_projects),
        ),
        {
            "hallucinated_projects": hallucinated_payload,
            "field_errors": field_errors,
        },
    )


_FIELD_PAIRS = (
    ("project_name", "project_name"),
    ("project_info", "project_info"),
    ("province", "province"),
    ("city", "city"),
    ("detailed_address", "detailed_address"),
    ("company_name", "company_name"),
    ("industry", "industry"),
    ("field", "field"),
    ("market", "market"),
)


def _calculate_dedupe_metrics(
    expected_dedupe: list[ExpectedDedupe],
    predictions: list[DedupePrediction],
) -> tuple[DedupeMetrics, dict[str, list[dict[str, object]]]]:
    predictions_by_key = {
        (_project_key(prediction.article_id, prediction.project_name)): prediction
        for prediction in predictions
    }
    decision_correct = 0
    status_change_correct = 0
    false_merges: list[dict[str, object]] = []
    missed_merges: list[dict[str, object]] = []

    for expected in expected_dedupe:
        prediction = predictions_by_key.get(
            _project_key(expected.article_id, expected.project_name)
        )
        predicted_decision = prediction.decision if prediction else "missing"
        if predicted_decision == expected.expected_decision:
            decision_correct += 1
        if (
            prediction is not None
            and prediction.status_change == expected.expected_status_change
        ):
            status_change_correct += 1

        if predicted_decision == "merge" and expected.expected_decision != "merge":
            false_merges.append(_dedupe_error(expected, predicted_decision))
        if expected.expected_decision == "merge" and predicted_decision != "merge":
            missed_merges.append(_dedupe_error(expected, predicted_decision))

    return (
        DedupeMetrics(
            dedupe_accuracy=_ratio(decision_correct, len(expected_dedupe)),
            false_merge_count=len(false_merges),
            missed_merge_count=len(missed_merges),
            status_change_accuracy=_ratio(
                status_change_correct,
                len(expected_dedupe),
            ),
        ),
        {
            "false_merges": false_merges,
            "missed_merges": missed_merges,
        },
    )


def _default_dedupe_decider(
    predictions: list[PredictedProject],
) -> list[DedupePrediction]:
    seen_by_name: dict[str, str] = {}
    last_status_by_name: dict[str, str | None] = {}
    dedupe_predictions: list[DedupePrediction] = []
    for prediction in predictions:
        project_name = prediction.project.project_name or ""
        key = normalize_project_name(project_name)
        if key and key in seen_by_name:
            decision = "merge"
            group_id = seen_by_name[key]
        else:
            decision = "new"
            group_id = key or None
            if key:
                seen_by_name[key] = key

        previous_status = last_status_by_name.get(key)
        status_change = bool(
            decision == "merge"
            and previous_status
            and prediction.project.status
            and prediction.project.status != previous_status
        )
        if key:
            last_status_by_name[key] = prediction.project.status

        dedupe_predictions.append(
            DedupePrediction(
                project_group_id=group_id,
                article_id=prediction.article_id,
                project_name=project_name,
                decision=decision,
                status_change=status_change,
            )
        )
    return dedupe_predictions


def _dedupe_error(expected: ExpectedDedupe, predicted_decision: str) -> dict[str, object]:
    return {
        "article_id": expected.article_id,
        "project_name": expected.project_name,
        "expected_decision": expected.expected_decision,
        "actual_decision": predicted_decision,
    }


def _project_key(article_id: str, project_name: str) -> tuple[str, str]:
    return article_id, normalize_project_name(project_name)


def _values_equal(left: object, right: object) -> bool:
    return _text(left) == _text(right)


def _numbers_close(left: float, right: float | None) -> bool:
    if right is None:
        return False
    return abs(float(left) - float(right)) <= 0.01


def _ratio(numerator: int, denominator: int) -> float:
    if denominator == 0:
        return 1.0
    return numerator / denominator


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _float(value: object) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _int(value: object) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def _bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = _text(value).lower()
    return text in {"1", "true", "yes", "y", "是"}
