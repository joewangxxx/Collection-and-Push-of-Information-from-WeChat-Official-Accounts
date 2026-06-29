from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from market_info.db.models import SourceArticle
from market_info.evaluation.core import (
    ARTICLE_HEADERS,
    EXPECTED_DEDUPE_HEADERS,
    EXPECTED_PROJECT_HEADERS,
)


def export_golden_template(
    session: Session,
    output_dir: Path,
    limit: int = 20,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    articles_dir = output_dir / "articles"
    articles_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    articles_sheet = workbook.active
    articles_sheet.title = "articles"
    expected_projects_sheet = workbook.create_sheet("expected_projects")
    expected_dedupe_sheet = workbook.create_sheet("expected_dedupe")

    articles_sheet.append(ARTICLE_HEADERS)
    expected_projects_sheet.append(EXPECTED_PROJECT_HEADERS)
    expected_dedupe_sheet.append(EXPECTED_DEDUPE_HEADERS)

    articles = (
        session.query(SourceArticle)
        .order_by(SourceArticle.published_at.desc(), SourceArticle.id.asc())
        .limit(limit)
        .all()
    )
    for article in articles:
        article_id = _stable_article_id(article)
        body_relative_path = Path("articles") / f"{article_id}.txt"
        body_path = output_dir / body_relative_path
        body_path.write_text(article.content_text or "", encoding="utf-8")
        articles_sheet.append(
            [
                article_id,
                article.account_name or "",
                article.title or "",
                _format_datetime(article.published_at),
                article.article_url or "",
                body_relative_path.as_posix(),
                "",
                "",
                "",
            ]
        )

    output_path = output_dir / "golden_labels.xlsx"
    workbook.save(output_path)
    return output_path


def _stable_article_id(article: SourceArticle) -> str:
    db_id = article.id if article.id is not None else "unknown"
    source = (
        article.normalized_url
        or article.article_url
        or f"{article.account_name}:{article.title}:{article.published_at}"
    )
    digest = hashlib.sha1(source.encode("utf-8")).hexdigest()[:12]
    return f"article_{db_id}_{digest}"


def _format_datetime(value: object) -> str:
    if value is None:
        return ""
    return str(value)
