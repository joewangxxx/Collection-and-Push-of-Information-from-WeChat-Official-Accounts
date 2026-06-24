from copy import copy
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from market_info.db.models import Project, ProjectEvent, ProjectRecord


UPDATED_SHEET = "本周新增与更新"
PROJECTS_SHEET = "项目全量台账"
REVIEW_SHEET = "疑似重复待复核"
SUMMARY_SHEET = "运行摘要"

UPDATED_HEADERS = [
    "发布日期",
    "公众号名称",
    "文章标题",
    "文章链接",
    "项目名称",
    "项目信息",
    "省份",
    "1级地级市",
    "详细地址",
    "企业名称",
    "项目投资额（亿）",
    "产业",
    "领域",
    "市场",
    "状态",
    "状态变化标注",
    "是否新增项目",
    "是否状态更新",
    "抽取置信度",
    "去重决策",
    "去重分数",
]

PROJECT_HEADERS = [
    "项目ID",
    "项目名称",
    "企业名称",
    "省份",
    "1级地级市",
    "详细地址",
    "项目投资额（亿）",
    "产业",
    "领域",
    "市场",
    "当前状态",
    "首次发现时间",
    "最近发现时间",
]

REVIEW_HEADERS = [
    "发布日期",
    "公众号名称",
    "文章标题",
    "文章链接",
    "项目名称",
    "企业名称",
    "省份",
    "1级地级市",
    "详细地址",
    "状态",
    "去重分数",
    "抽取置信度",
]

SUMMARY_HEADERS = ["指标", "数值"]


def generate_weekly_excel(
    session: Session,
    output_dir: Path,
    run_id: str | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    filename_id = run_id or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"market_info_weekly_{filename_id}.xlsx"

    workbook = Workbook()
    updated_sheet = workbook.active
    updated_sheet.title = UPDATED_SHEET
    projects_sheet = workbook.create_sheet(PROJECTS_SHEET)
    review_sheet = workbook.create_sheet(REVIEW_SHEET)
    summary_sheet = workbook.create_sheet(SUMMARY_SHEET)

    events_by_record_key = _load_events_by_record_key(session)
    updated_records = (
        session.query(ProjectRecord)
        .filter(ProjectRecord.dedupe_decision.in_(("new", "merge")))
        .order_by(ProjectRecord.id)
        .all()
    )
    review_records = (
        session.query(ProjectRecord)
        .filter(ProjectRecord.dedupe_decision == "review")
        .order_by(ProjectRecord.id)
        .all()
    )
    projects = session.query(Project).order_by(Project.id).all()

    _write_sheet(
        updated_sheet,
        UPDATED_HEADERS,
        [
            _updated_record_row(record, events_by_record_key)
            for record in updated_records
        ],
    )
    _write_sheet(
        projects_sheet,
        PROJECT_HEADERS,
        [_project_row(project) for project in projects],
    )
    _write_sheet(
        review_sheet,
        REVIEW_HEADERS,
        [_review_record_row(record) for record in review_records],
    )
    _write_sheet(
        summary_sheet,
        SUMMARY_HEADERS,
        _summary_rows(session, generated_at=datetime.now()),
    )

    workbook.save(output_path)
    return output_path


def _updated_record_row(
    record: ProjectRecord,
    events_by_record_key: dict[tuple[int, int], ProjectEvent],
) -> list[object]:
    article = record.source_article
    event = _record_event(record, events_by_record_key)
    return [
        _cell_value(article.published_at),
        article.account_name,
        article.title,
        article.article_url,
        record.project_name,
        record.project_info,
        record.province,
        record.city,
        record.detailed_address,
        record.company_name,
        record.investment_amount_yi,
        record.industry,
        record.field,
        record.market,
        record.status,
        event.change_label if event else "",
        "是" if record.dedupe_decision == "new" else "否",
        "是" if event else "否",
        record.confidence,
        record.dedupe_decision,
        record.dedupe_score,
    ]


def _project_row(project: Project) -> list[object]:
    return [
        project.id,
        project.canonical_project_name,
        project.canonical_company_name,
        project.province,
        project.city,
        project.detailed_address,
        project.investment_amount_yi,
        project.industry,
        project.field,
        project.market,
        project.current_status,
        _cell_value(project.first_seen_at),
        _cell_value(project.last_seen_at),
    ]


def _review_record_row(record: ProjectRecord) -> list[object]:
    article = record.source_article
    return [
        _cell_value(article.published_at),
        article.account_name,
        article.title,
        article.article_url,
        record.project_name,
        record.company_name,
        record.province,
        record.city,
        record.detailed_address,
        record.status,
        record.dedupe_score,
        record.confidence,
    ]


def _summary_rows(session: Session, generated_at: datetime) -> list[list[object]]:
    return [
        ["生成时间", _cell_value(generated_at)],
        [
            "新增项目记录数",
            session.query(ProjectRecord)
            .filter(ProjectRecord.dedupe_decision == "new")
            .count(),
        ],
        [
            "合并项目记录数",
            session.query(ProjectRecord)
            .filter(ProjectRecord.dedupe_decision == "merge")
            .count(),
        ],
        [
            "疑似重复待复核数",
            session.query(ProjectRecord)
            .filter(ProjectRecord.dedupe_decision == "review")
            .count(),
        ],
        ["项目台账总数", session.query(Project).count()],
        ["状态变化事件数", session.query(ProjectEvent).count()],
    ]


def _load_events_by_record_key(session: Session) -> dict[tuple[int, int], ProjectEvent]:
    events: dict[tuple[int, int], ProjectEvent] = {}
    for event in session.query(ProjectEvent).order_by(ProjectEvent.id).all():
        key = (event.project_id, event.source_article_id)
        events.setdefault(key, event)
    return events


def _record_event(
    record: ProjectRecord,
    events_by_record_key: dict[tuple[int, int], ProjectEvent],
) -> ProjectEvent | None:
    if record.project_id is None or record.source_article_id is None:
        return None
    return events_by_record_key.get((record.project_id, record.source_article_id))


def _write_sheet(
    sheet: Worksheet,
    headers: list[str],
    rows: list[list[object]],
) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append([_cell_value(value) for value in row])

    _autosize_columns(sheet)


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def _cell_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    return value
