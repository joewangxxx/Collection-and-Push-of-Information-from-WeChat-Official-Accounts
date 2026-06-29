import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from market_info.ai.schemas import ExtractedProject
from market_info.db.base import Base
from market_info.db.models import SourceArticle
from market_info.evaluation import (
    DedupePrediction,
    ExtractorProtocol,
    PredictedProject,
    evaluate_golden,
    export_golden_template,
    load_golden_labels,
)
from market_info.evaluation.core import ARTICLE_HEADERS


class FakeExtractor:
    def __init__(self, by_title: dict[str, list[ExtractedProject]]) -> None:
        self.by_title = by_title
        self.calls: list[tuple[str, str]] = []

    def extract(self, title: str, text: str) -> list[ExtractedProject]:
        self.calls.append((title, text))
        return self.by_title.get(title, [])


def write_labels_fixture(tmp_path: Path) -> Path:
    articles_dir = tmp_path / "articles"
    articles_dir.mkdir()
    (articles_dir / "a1.txt").write_text("正文：项目备案，投资10亿元。", encoding="utf-8")
    (articles_dir / "a2.txt").write_text("正文：行业新闻，无项目。", encoding="utf-8")
    (articles_dir / "a3.txt").write_text("正文：项目投产，投资12亿元。", encoding="utf-8")

    workbook = Workbook()
    articles_sheet = workbook.active
    articles_sheet.title = "articles"
    articles_sheet.append(
        [
            "article_id",
            "account_name",
            "title",
            "published_at",
            "url",
            "body_path",
            "is_project_article",
            "expected_project_count",
            "notes",
        ]
    )
    articles_sheet.append(
        ["a1", "光伏前沿", "A项目备案", "2026-06-01", "https://mp.weixin.qq.com/s/a1", "articles/a1.txt", True, 1, ""]
    )
    articles_sheet.append(
        ["a2", "光伏前沿", "行业新闻", "2026-06-02", "https://mp.weixin.qq.com/s/a2", "articles/a2.txt", False, 0, ""]
    )
    articles_sheet.append(
        ["a3", "光伏前沿", "A项目投产", "2026-06-03", "https://mp.weixin.qq.com/s/a3", "articles/a3.txt", True, 1, ""]
    )

    expected_projects = workbook.create_sheet("expected_projects")
    expected_projects.append(
        [
            "article_id",
            "project_group_id",
            "project_name",
            "project_info",
            "province",
            "city",
            "detailed_address",
            "company_name",
            "investment_yi",
            "industry",
            "field",
            "market",
            "status",
            "evidence",
            "notes",
        ]
    )
    expected_projects.append(
        ["a1", "g1", "A光伏项目", "建设光伏基地", "江苏省", "盐城市", "经开区", "A公司", 10, "新能源", "光伏", "电力", "备案", "投资10亿元", ""]
    )
    expected_projects.append(
        ["a3", "g1", "A光伏项目", "建设光伏基地", "江苏省", "盐城市", "经开区", "A公司", 12, "新能源", "光伏", "电力", "投产", "项目投产", ""]
    )

    expected_dedupe = workbook.create_sheet("expected_dedupe")
    expected_dedupe.append(
        [
            "project_group_id",
            "article_id",
            "project_name",
            "expected_decision",
            "expected_status_change",
            "status_order",
            "notes",
        ]
    )
    expected_dedupe.append(["g1", "a1", "A光伏项目", "new", False, 1, ""])
    expected_dedupe.append(["g1", "a3", "A光伏项目", "merge", True, 2, ""])

    labels_path = tmp_path / "golden_labels.xlsx"
    workbook.save(labels_path)
    return labels_path


def test_load_golden_labels_reads_articles_projects_and_dedupe(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)

    labels = load_golden_labels(labels_path)

    assert [article.article_id for article in labels.articles] == ["a1", "a2", "a3"]
    assert labels.articles[0].body_text == "正文：项目备案，投资10亿元。"
    assert labels.expected_projects[0].project_group_id == "g1"
    assert labels.expected_dedupe[1].expected_decision == "merge"


def test_extractor_protocol_is_exported_from_public_api() -> None:
    assert ExtractorProtocol.__name__ == "ExtractorProtocol"


def test_load_golden_labels_rejects_missing_sheet(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)
    workbook = load_workbook(labels_path)
    del workbook["expected_dedupe"]
    workbook.save(labels_path)

    try:
        load_golden_labels(labels_path)
    except ValueError as exc:
        assert "expected_dedupe" in str(exc)
        assert "missing" in str(exc)
    else:
        raise AssertionError("missing sheet should raise ValueError")


def test_load_golden_labels_rejects_wrong_header_order(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)
    workbook = load_workbook(labels_path)
    workbook["articles"].delete_cols(1)
    workbook["articles"].insert_cols(2)
    workbook["articles"].cell(row=1, column=2).value = ARTICLE_HEADERS[0]
    workbook.save(labels_path)

    try:
        load_golden_labels(labels_path)
    except ValueError as exc:
        assert "articles" in str(exc)
        assert "expected headers" in str(exc)
    else:
        raise AssertionError("invalid header should raise ValueError")


def test_load_golden_labels_rejects_expected_project_count_mismatch(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)
    workbook = load_workbook(labels_path)
    workbook["articles"].cell(row=2, column=8).value = 2
    workbook.save(labels_path)

    try:
        load_golden_labels(labels_path)
    except ValueError as exc:
        assert "a1" in str(exc)
        assert "expected_project_count" in str(exc)
    else:
        raise AssertionError("count mismatch should raise ValueError")


def test_load_golden_labels_rejects_non_project_article_with_expected_projects(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)
    workbook = load_workbook(labels_path)
    workbook["expected_projects"].append(
        ["a2", "g2", "不应存在项目", "", "", "", "", "", "", "", "", "", "", "", ""]
    )
    workbook.save(labels_path)

    try:
        load_golden_labels(labels_path)
    except ValueError as exc:
        assert "a2" in str(exc)
        assert "is_project_article is False" in str(exc)
    else:
        raise AssertionError("non-project article with project rows should raise ValueError")


def test_evaluate_golden_computes_extraction_and_dedupe_metrics(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)
    extractor = FakeExtractor(
        {
            "A项目备案": [
                ExtractedProject(
                    project_name="A光伏项目",
                    project_info="建设光伏基地",
                    province="江苏省",
                    city="盐城市",
                    detailed_address="经开区",
                    company_name="A公司",
                    investment_amount_yi=10,
                    industry="新能源",
                    field="光伏",
                    market="电力",
                    status="备案",
                    confidence=0.9,
                )
            ],
            "行业新闻": [
                ExtractedProject(
                    project_name="幻觉项目",
                    status="未知",
                    confidence=0.4,
                )
            ],
            "A项目投产": [],
        }
    )

    def fake_dedupe(predictions: list[PredictedProject]) -> list[DedupePrediction]:
        return [
            DedupePrediction(
                project_group_id="g1",
                article_id="a1",
                project_name="A光伏项目",
                decision="new",
                status_change=False,
            )
        ]

    report = evaluate_golden(
        labels_path,
        extractor=extractor,
        dedupe_decider=fake_dedupe,
    )

    assert report.extraction.project_precision == 0.5
    assert report.extraction.project_recall == 0.5
    assert report.extraction.hallucination_count == 1
    assert report.extraction.missed_count == 1
    assert report.extraction.status_accuracy == 1.0
    assert report.dedupe.dedupe_accuracy == 0.5
    assert report.dedupe.missed_merge_count == 1
    assert report.error_samples.missed_articles == ["a3"]
    assert report.error_samples.hallucinated_projects[0]["article_id"] == "a2"


def test_status_accuracy_only_counts_projects_with_status_labels(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)
    workbook = load_workbook(labels_path)
    workbook["expected_projects"].cell(row=3, column=13).value = ""
    workbook.save(labels_path)
    extractor = FakeExtractor(
        {
            "A项目备案": [
                ExtractedProject(
                    project_name="A光伏项目",
                    status="备案",
                    confidence=0.9,
                )
            ],
            "A项目投产": [
                ExtractedProject(
                    project_name="A光伏项目",
                    status="未知",
                    confidence=0.9,
                )
            ],
        }
    )

    report = evaluate_golden(labels_path, extractor=extractor)

    assert report.extraction.status_accuracy == 1.0


def test_evaluate_golden_writes_json_report(tmp_path) -> None:
    labels_path = write_labels_fixture(tmp_path)
    report_path = tmp_path / "evaluation_report.json"
    extractor = FakeExtractor({"A项目备案": [], "行业新闻": [], "A项目投产": []})

    evaluate_golden(labels_path, extractor=extractor, report_path=report_path)

    payload = json.loads(report_path.read_text(encoding="utf-8"))
    assert "extraction" in payload
    assert "dedupe" in payload
    assert "error_samples" in payload


def test_export_golden_template_writes_articles_and_label_workbook(tmp_path) -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    session = SessionLocal()
    try:
        session.add(
            SourceArticle(
                account_id=1,
                account_name="光伏前沿",
                title="测试文章",
                article_url="https://mp.weixin.qq.com/s/export",
                normalized_url="https://mp.weixin.qq.com/s/export",
                published_at=datetime(2026, 6, 1, tzinfo=timezone.utc),
                content_text="文章正文",
                content_hash="a" * 64,
            )
        )
        session.flush()

        labels_path = export_golden_template(session, tmp_path, limit=1)
    finally:
        session.close()

    workbook = load_workbook(labels_path)
    assert workbook.sheetnames == ["articles", "expected_projects", "expected_dedupe"]
    article_row = list(workbook["articles"].iter_rows(min_row=2, values_only=True))[0]
    assert article_row[1] == "光伏前沿"
    body_path = tmp_path / article_row[5]
    assert body_path.read_text(encoding="utf-8") == "文章正文"


def test_export_golden_template_uses_database_id_to_avoid_duplicate_url_overwrite(tmp_path) -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    session = SessionLocal()
    try:
        for index, body in enumerate(("第一篇正文", "第二篇正文"), start=1):
            session.add(
                SourceArticle(
                    account_id=1,
                    account_name="光伏前沿",
                    title="重复链接文章",
                    article_url="https://mp.weixin.qq.com/s/same",
                    normalized_url=f"https://mp.weixin.qq.com/s/same?idx={index}",
                    published_at=datetime(2026, 6, 1, tzinfo=timezone.utc),
                    content_text=body,
                    content_hash=body[:1] * 64,
                )
            )
        session.flush()
        labels_path = export_golden_template(session, tmp_path, limit=2)
    finally:
        session.close()

    workbook = load_workbook(labels_path)
    rows = list(workbook["articles"].iter_rows(min_row=2, values_only=True))
    article_ids = [row[0] for row in rows]
    body_paths = [row[5] for row in rows]
    assert len(set(article_ids)) == 2
    assert len(set(body_paths)) == 2
    assert sorted((tmp_path / path).read_text(encoding="utf-8") for path in body_paths) == [
        "第一篇正文",
        "第二篇正文",
    ]
