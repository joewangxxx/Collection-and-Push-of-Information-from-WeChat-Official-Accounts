from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from market_info.db.base import Base
from market_info.db.models import Project, ProjectEvent, ProjectRecord, SourceArticle
from market_info.reports.excel_report import generate_weekly_excel


SHEET_UPDATED = "本周新增与更新"
SHEET_PROJECTS = "项目全量台账"
SHEET_REVIEW = "疑似重复待复核"
SHEET_SUMMARY = "运行摘要"


@pytest.fixture()
def session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def make_article(
    title: str,
    normalized_url: str,
    published_at: datetime | None,
) -> SourceArticle:
    return SourceArticle(
        account_id=1,
        account_name="光伏前沿",
        title=title,
        article_url=f"https://mp.weixin.qq.com/s/{normalized_url}",
        normalized_url=f"https://mp.weixin.qq.com/s/{normalized_url}",
        published_at=published_at,
        content_text="正文",
        content_hash=(normalized_url[:1] or "x") * 64,
    )


def make_project(name: str, status: str) -> Project:
    return Project(
        canonical_project_name=name,
        canonical_company_name="江苏示例新能源有限公司",
        province="江苏省",
        city="盐城市",
        detailed_address="盐城经开区",
        investment_amount_yi=10.5,
        industry="新能源",
        field="光伏",
        market="电力",
        current_status=status,
        first_seen_at=datetime(2026, 6, 1, 8, 0, tzinfo=timezone.utc),
        last_seen_at=datetime(2026, 6, 8, 8, 0, tzinfo=timezone.utc),
        semantic_text="项目语义文本",
        embedding=None,
    )


def make_record(
    article: SourceArticle,
    project: Project | None,
    name: str,
    decision: str,
    score: float,
    status: str,
) -> ProjectRecord:
    return ProjectRecord(
        source_article=article,
        project=project,
        project_name=name,
        project_info=None if decision == "review" else "新增组件产线",
        province="江苏省",
        city="盐城市",
        detailed_address=None if decision == "review" else "盐城经开区",
        company_name="江苏示例新能源有限公司",
        investment_amount_yi=10.5,
        industry="新能源",
        field="光伏",
        market="电力",
        status=status,
        confidence=0.91,
        semantic_text="项目语义文本",
        embedding=None,
        dedupe_decision=decision,
        dedupe_score=score,
    )


def seed_report_data(session) -> None:
    new_article = make_article(
        "新增项目文章",
        "new",
        datetime(2026, 6, 2, 8, 0, tzinfo=timezone.utc),
    )
    merge_article = make_article(
        "状态更新文章",
        "merge",
        datetime(2026, 6, 3, 8, 0, tzinfo=timezone.utc),
    )
    review_article = make_article(
        "疑似重复文章",
        "review",
        datetime(2026, 6, 4, 8, 0, tzinfo=timezone.utc),
    )
    new_project = make_project("新增光伏组件项目", "备案")
    existing_project = make_project("已有光伏项目", "开工")
    new_record = make_record(new_article, new_project, "新增光伏组件项目", "new", 92.5, "备案")
    merge_record = make_record(merge_article, existing_project, "已有光伏项目", "merge", 88.2, "开工")
    review_record = make_record(review_article, None, "疑似重复光伏项目", "review", 76.4, "备案")
    session.add_all([new_record, merge_record, review_record])
    session.flush()
    session.add(
        ProjectEvent(
            project=existing_project,
            source_article=merge_article,
            event_status="开工",
            previous_status="备案",
            event_date=merge_article.published_at,
            change_label="备案 -> 开工",
        )
    )
    session.flush()


def load_generated_workbook(session, tmp_path):
    seed_report_data(session)
    output_path = generate_weekly_excel(session, tmp_path / "nested", run_id="test-run")
    workbook = load_workbook(output_path)
    return output_path, workbook


def rows_as_dicts(sheet):
    headers = [cell.value for cell in sheet[1]]
    return [
        dict(zip(headers, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in row)
    ]


def test_generate_weekly_excel_creates_xlsx_file(session, tmp_path) -> None:
    output_path, _ = load_generated_workbook(session, tmp_path)

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"
    assert output_path.name == "market_info_weekly_test-run.xlsx"


def test_workbook_contains_required_sheets(session, tmp_path) -> None:
    _, workbook = load_generated_workbook(session, tmp_path)

    assert workbook.sheetnames == [SHEET_UPDATED, SHEET_PROJECTS, SHEET_REVIEW, SHEET_SUMMARY]


def test_updated_sheet_contains_required_headers_and_formatting(session, tmp_path) -> None:
    _, workbook = load_generated_workbook(session, tmp_path)
    sheet = workbook[SHEET_UPDATED]
    headers = [cell.value for cell in sheet[1]]

    assert headers == [
        "发布日期",
        "公众号名称",
        "文章标题",
        "文章链接",
        "项目名称",
        "项目信息",
        "省份",
        "1级地级市",
        "详细地址",
        "企业名称",
        "项目投资额（亿）",
        "产业",
        "领域",
        "市场",
        "状态",
        "状态变化标注",
        "是否新增项目",
        "是否状态更新",
        "抽取置信度",
        "去重决策",
        "去重分数",
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].font.bold is True
    assert sheet.column_dimensions["A"].width > 8


def test_new_and_merge_records_enter_updated_sheet_with_flags(session, tmp_path) -> None:
    _, workbook = load_generated_workbook(session, tmp_path)
    rows = rows_as_dicts(workbook[SHEET_UPDATED])

    assert [row["项目名称"] for row in rows] == ["新增光伏组件项目", "已有光伏项目"]
    new_row = rows[0]
    merge_row = rows[1]
    assert new_row["公众号名称"] == "光伏前沿"
    assert new_row["文章链接"] == "https://mp.weixin.qq.com/s/new"
    assert new_row["是否新增项目"] == "是"
    assert new_row["是否状态更新"] == "否"
    assert merge_row["是否新增项目"] == "否"
    assert merge_row["是否状态更新"] == "是"
    assert merge_row["状态变化标注"] == "备案 -> 开工"


def test_review_record_enters_review_sheet_only(session, tmp_path) -> None:
    _, workbook = load_generated_workbook(session, tmp_path)
    updated_names = {row["项目名称"] for row in rows_as_dicts(workbook[SHEET_UPDATED])}
    review_rows = rows_as_dicts(workbook[SHEET_REVIEW])

    assert "疑似重复光伏项目" not in updated_names
    assert len(review_rows) == 1
    assert review_rows[0]["项目名称"] == "疑似重复光伏项目"
    assert review_rows[0]["去重分数"] == 76.4


def test_project_sheet_contains_existing_projects(session, tmp_path) -> None:
    _, workbook = load_generated_workbook(session, tmp_path)
    rows = rows_as_dicts(workbook[SHEET_PROJECTS])
    names = {row["项目名称"] for row in rows}

    assert {"新增光伏组件项目", "已有光伏项目"} <= names
    assert {row["当前状态"] for row in rows} >= {"备案", "开工"}


def test_summary_sheet_counts_are_correct(session, tmp_path) -> None:
    _, workbook = load_generated_workbook(session, tmp_path)
    summary = {row[0]: row[1] for row in workbook[SHEET_SUMMARY].iter_rows(min_row=2, values_only=True)}

    assert summary["新增项目记录数"] == 1
    assert summary["合并项目记录数"] == 1
    assert summary["疑似重复待复核数"] == 1
    assert summary["项目台账总数"] == 2
    assert summary["状态变化事件数"] == 1
    assert summary["生成时间"]


def test_empty_fields_are_written_as_empty_string_not_none_text(session, tmp_path) -> None:
    _, workbook = load_generated_workbook(session, tmp_path)
    review_rows = rows_as_dicts(workbook[SHEET_REVIEW])

    assert review_rows[0]["详细地址"] is None
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            assert "None" not in [value for value in row if isinstance(value, str)]


def test_generate_weekly_excel_does_not_commit(session, tmp_path) -> None:
    seed_report_data(session)

    def fail_commit() -> None:
        raise AssertionError("generate_weekly_excel must not commit")

    session.commit = fail_commit

    generate_weekly_excel(session, tmp_path, run_id="no-commit")
